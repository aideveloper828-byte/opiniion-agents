#!/usr/bin/env python3
"""
Monthly Operating Rhythm — Opiniion
Orchestrates close checklist tracking, BvA variance, Finance Flash,
and mid-month re-forecast deliverables.

Usage:
    python3 run_close_cycle.py --phase close --month 2026-04
    python3 run_close_cycle.py --phase post-close --month 2026-04
    python3 run_close_cycle.py --phase reforecast --month 2026-04
    python3 run_close_cycle.py --phase all --month 2026-04
"""

import argparse
import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from connectors.netsuite_connect import NetSuiteClient

TEAL = "4FCBC5"
DARK_GRAY = "4C4D4E"
LIGHT_TEAL = "E0F5F4"
HEADER_GRAY = "F3F3F3"
WHITE = "FFFFFF"
RAG_RED = "FF6B6B"
RAG_AMBER = "FFD93D"
RAG_GREEN = "6BCB77"
VAR_RED = "FCE4D6"
VAR_GREEN = "E2EFDA"

BRAND_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
TEAL_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
SECTION_FILL = PatternFill(start_color=HEADER_GRAY, end_color=HEADER_GRAY, fill_type="solid")
VALUE_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_GRAY)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
CURRENCY_FMT = '"$"#,##0'
PCT_FMT = "0.0%"

VARIANCE_THRESHOLD_PCT = 0.10
VARIANCE_THRESHOLD_ABS = 5000


CLOSE_CHECKLIST = [
    {"phase": "Pre-Close", "task": "Confirm all invoices issued", "due": "EOM - 2"},
    {"phase": "Pre-Close", "task": "Verify deferred revenue schedule", "due": "EOM - 1"},
    {"phase": "Pre-Close", "task": "Confirm payroll posted", "due": "EOM - 1"},
    {"phase": "Pre-Close", "task": "Post manual journal entries", "due": "EOM"},
    {"phase": "Close", "task": "Bank reconciliation", "due": "BD +1"},
    {"phase": "Close", "task": "AP cutoff review", "due": "BD +2"},
    {"phase": "Close", "task": "Revenue recognition review", "due": "BD +2"},
    {"phase": "Close", "task": "Intercompany reconciliation", "due": "BD +2"},
    {"phase": "Close", "task": "Prepaid amortization", "due": "BD +2"},
    {"phase": "Close", "task": "Fixed asset depreciation", "due": "BD +2"},
    {"phase": "Close", "task": "Balance sheet reconciliation", "due": "BD +3"},
    {"phase": "Close", "task": "Trial balance review", "due": "BD +3"},
    {"phase": "Post-Close", "task": "Close period in NetSuite", "due": "BD +3"},
    {"phase": "Post-Close", "task": "Generate financial statements", "due": "BD +3"},
    {"phase": "Post-Close", "task": "Run SaaS metrics", "due": "BD +4"},
    {"phase": "Post-Close", "task": "BvA variance analysis", "due": "BD +4"},
    {"phase": "Post-Close", "task": "Build Finance Flash", "due": "BD +4"},
    {"phase": "Post-Close", "task": "Distribute to CEO/COO", "due": "BD +5"},
    {"phase": "Post-Close", "task": "Archive close package", "due": "BD +5"},
]


def pull_trial_balance(client: NetSuiteClient, month: str) -> pd.DataFrame:
    """Pull trial balance for the specified month."""
    start_date = f"{month}-01"
    from dateutil.relativedelta import relativedelta
    end_dt = datetime.strptime(start_date, "%Y-%m-%d") + relativedelta(months=1)
    end_date = end_dt.strftime("%Y-%m-%d")

    sql = f"""
    SELECT
        a.acctnumber AS account_number,
        a.acctname AS account_name,
        a.accttype AS account_type,
        SUM(tal.amount) AS balance
    FROM transactionaccountingline tal
    JOIN transaction t ON tal.transaction = t.id
    JOIN account a ON tal.account = a.id
    JOIN accountingperiod ap ON tal.postingperiod = ap.id
    WHERE ap.startdate >= '{start_date}'
      AND ap.startdate < '{end_date}'
      AND ap.isadjust = 'F'
      AND ap.isquarter = 'F'
      AND ap.isyear = 'F'
    GROUP BY a.acctnumber, a.acctname, a.accttype
    HAVING SUM(tal.amount) != 0
    ORDER BY a.acctnumber
    """
    print(f"[Close] Pulling trial balance for {month}...")
    df = client.query(sql)
    if not df.empty:
        df["balance"] = pd.to_numeric(df["balance"], errors="coerce").fillna(0)
    return df


# ── Close Checklist Workbook ───────────────────────────────────────

def build_close_tracker(month: str, output_path: str):
    """Build the close checklist tracker Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Close Checklist"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    ws.cell(row=2, column=1, value=f"Monthly Close — {month}").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )

    headers = ["#", "Phase", "Task", "Due", "Status", "RAG", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(CLOSE_CHECKLIST, 5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4).font = VALUE_FONT
        ws.cell(row=row_idx, column=2, value=item["phase"]).font = VALUE_FONT
        ws.cell(row=row_idx, column=3, value=item["task"]).font = VALUE_FONT
        ws.cell(row=row_idx, column=4, value=item["due"]).font = VALUE_FONT
        ws.cell(row=row_idx, column=5, value="Not Started").font = VALUE_FONT
        ws.cell(row=row_idx, column=6, value="Green").font = VALUE_FONT
        rag_fill = PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid")
        ws.cell(row=row_idx, column=6).fill = rag_fill
        ws.cell(row=row_idx, column=7, value="").font = VALUE_FONT

    col_widths = [5, 12, 35, 10, 14, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"[Close] Checklist saved to: {output_path}")


# ── BvA Variance Report ───────────────────────────────────────────

def build_bva_report(
    actuals: pd.DataFrame,
    budget: pd.DataFrame,
    month: str,
    output_path: str,
):
    """Build Budget vs. Actuals variance report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BvA Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    ws.cell(row=2, column=1, value=f"Budget vs. Actuals — {month}").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )

    if actuals.empty:
        ws.cell(row=4, column=1, value="No actuals data available for this period.").font = VALUE_FONT
        wb.save(output_path)
        return

    headers = ["Account", "Actual", "Budget", "Variance ($)", "Variance (%)", "Flag"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    if budget.empty:
        for row_idx, (_, row_data) in enumerate(actuals.iterrows(), 5):
            ws.cell(row=row_idx, column=1, value=f"{row_data.get('account_number', '')} - {row_data.get('account_name', '')}").font = VALUE_FONT
            ws.cell(row=row_idx, column=2, value=row_data.get("balance", 0)).font = VALUE_FONT
            ws.cell(row=row_idx, column=2).number_format = CURRENCY_FMT
            ws.cell(row=row_idx, column=3, value="No budget").font = Font(name="Calibri", size=10, italic=True, color="999999")
    else:
        merged = actuals.merge(budget, on="account_number", how="outer", suffixes=("_actual", "_budget"))
        merged["actual"] = pd.to_numeric(merged.get("balance_actual", 0), errors="coerce").fillna(0)
        merged["budget"] = pd.to_numeric(merged.get("balance_budget", 0), errors="coerce").fillna(0)
        merged["variance_dollar"] = merged["actual"] - merged["budget"]
        merged["variance_pct"] = merged.apply(
            lambda r: r["variance_dollar"] / r["budget"] if r["budget"] != 0 else 0, axis=1
        )

        for row_idx, (_, row_data) in enumerate(merged.iterrows(), 5):
            acct = f"{row_data.get('account_number', '')} - {row_data.get('account_name_actual', row_data.get('account_name_budget', ''))}"
            ws.cell(row=row_idx, column=1, value=acct).font = VALUE_FONT
            ws.cell(row=row_idx, column=2, value=row_data["actual"]).number_format = CURRENCY_FMT
            ws.cell(row=row_idx, column=2).font = VALUE_FONT
            ws.cell(row=row_idx, column=3, value=row_data["budget"]).number_format = CURRENCY_FMT
            ws.cell(row=row_idx, column=3).font = VALUE_FONT
            ws.cell(row=row_idx, column=4, value=row_data["variance_dollar"]).number_format = CURRENCY_FMT
            ws.cell(row=row_idx, column=4).font = VALUE_FONT
            ws.cell(row=row_idx, column=5, value=row_data["variance_pct"]).number_format = PCT_FMT
            ws.cell(row=row_idx, column=5).font = VALUE_FONT

            is_material = (
                abs(row_data["variance_pct"]) > VARIANCE_THRESHOLD_PCT
                and abs(row_data["variance_dollar"]) > VARIANCE_THRESHOLD_ABS
            )
            if is_material:
                flag = "OVER" if row_data["variance_dollar"] > 0 else "UNDER"
                ws.cell(row=row_idx, column=6, value=flag).font = Font(
                    name="Calibri", size=10, bold=True,
                    color="CC0000" if flag == "OVER" else "006600"
                )
                fill_color = VAR_RED if flag == "OVER" else VAR_GREEN
                for c in range(1, 7):
                    ws.cell(row=row_idx, column=c).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

    col_widths = [40, 14, 14, 16, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"[Close] BvA report saved to: {output_path}")


# ── Finance Flash ──────────────────────────────────────────────────

def build_finance_flash(
    actuals: pd.DataFrame,
    month: str,
    output_path: str,
):
    """Build the 1-page Finance Flash executive summary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Flash"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    ws.cell(row=2, column=1, value=f"Finance Flash — {month}").font = Font(
        name="Calibri", size=16, bold=True, color=DARK_GRAY
    )
    ws.cell(row=2, column=3, value=f"Prepared: {datetime.today().strftime('%Y-%m-%d')}").font = Font(
        name="Calibri", size=9, color="999999"
    )

    row = 4
    sections = [
        ("Financial Performance", [
            ("Total Revenue", "Populate from financial-statement-builder", CURRENCY_FMT),
            ("Gross Margin %", "Populate from financial-statement-builder", PCT_FMT),
            ("Total OpEx", "Populate from financial-statement-builder", CURRENCY_FMT),
            ("EBITDA", "Populate from financial-statement-builder", CURRENCY_FMT),
            ("Cash Balance", "Populate from NetSuite", CURRENCY_FMT),
        ]),
        ("SaaS Metrics", [
            ("Ending ARR", "Populate from saas-metrics-engine", CURRENCY_FMT),
            ("Net Retention (NRR)", "Populate from saas-metrics-engine", PCT_FMT),
            ("Gross Retention (GRR)", "Populate from saas-metrics-engine", PCT_FMT),
            ("Magic Number", "Populate from saas-metrics-engine", "0.0x"),
        ]),
        ("Pipeline & Sales", [
            ("Total Pipeline", "Populate from pipeline-intelligence", CURRENCY_FMT),
            ("Weighted Pipeline", "Populate from pipeline-intelligence", CURRENCY_FMT),
            ("Pipeline Coverage", "Populate from pipeline-intelligence", "0.0x"),
            ("Win Rate", "Populate from pipeline-intelligence", PCT_FMT),
        ]),
    ]

    for section_name, items in sections:
        ws.cell(row=row, column=1, value=section_name).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=2).fill = SECTION_FILL
        ws.cell(row=row, column=3).fill = SECTION_FILL
        row += 1
        for label, placeholder, fmt in items:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            cell = ws.cell(row=row, column=2, value=placeholder)
            cell.font = Font(name="Calibri", size=10, italic=True, color="999999")
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Key Callouts").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=2).fill = SECTION_FILL
    ws.cell(row=row, column=3).fill = SECTION_FILL
    row += 1
    ws.cell(row=row, column=1, value="1.").font = VALUE_FONT
    ws.cell(row=row, column=2, value="[Key observation about financial performance]").font = Font(
        name="Calibri", size=10, italic=True, color="999999"
    )
    row += 1
    ws.cell(row=row, column=1, value="2.").font = VALUE_FONT
    ws.cell(row=row, column=2, value="[Key observation about pipeline or growth]").font = Font(
        name="Calibri", size=10, italic=True, color="999999"
    )
    row += 1
    ws.cell(row=row, column=1, value="3.").font = VALUE_FONT
    ws.cell(row=row, column=2, value="[Risk or action item requiring attention]").font = Font(
        name="Calibri", size=10, italic=True, color="999999"
    )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20

    wb.save(output_path)
    print(f"[Close] Finance Flash saved to: {output_path}")


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Opiniion Monthly Operating Rhythm")
    parser.add_argument("--phase", choices=["close", "post-close", "reforecast", "all"], default="all")
    parser.add_argument("--month", default=datetime.today().strftime("%Y-%m"), help="Target month (YYYY-MM)")
    parser.add_argument("--output-dir", default=None, help="Output directory")
    args = parser.parse_args()

    output_dir = args.output_dir or os.path.expanduser(f"~/Desktop/Opiniion_Close_{args.month}")
    os.makedirs(output_dir, exist_ok=True)

    client = NetSuiteClient()

    if args.phase in ("close", "all"):
        print(f"\n{'='*60}")
        print(f"PHASE: CLOSE WINDOW — {args.month}")
        print(f"{'='*60}")
        build_close_tracker(args.month, os.path.join(output_dir, f"Close_Checklist_{args.month}.xlsx"))

        print("[Close] Pulling trial balance for soft close review...")
        tb = pull_trial_balance(client, args.month)
        if not tb.empty:
            print(f"[Close] Trial balance: {len(tb)} accounts with activity")
        else:
            print("[Close] No trial balance data — may not be closed yet")

    if args.phase in ("post-close", "all"):
        print(f"\n{'='*60}")
        print(f"PHASE: POST-CLOSE — {args.month}")
        print(f"{'='*60}")

        actuals = pull_trial_balance(client, args.month)
        budget = pd.DataFrame()  # placeholder until budget source is configured

        build_bva_report(
            actuals, budget, args.month,
            os.path.join(output_dir, f"BvA_{args.month}.xlsx"),
        )

        build_finance_flash(
            actuals, args.month,
            os.path.join(output_dir, f"Finance_Flash_{args.month}.xlsx"),
        )

        print("[Close] NOTE: Run saas-metrics-engine and pipeline-intelligence")
        print("[Close]       separately to populate Finance Flash with live data.")

    if args.phase in ("reforecast", "all"):
        print(f"\n{'='*60}")
        print(f"PHASE: MID-MONTH RE-FORECAST — {args.month}")
        print(f"{'='*60}")
        print("[Reforecast] Run pipeline-intelligence for pipeline snapshot.")
        print("[Reforecast] Run financial-statement-builder with --include-forecast for updated projections.")
        print("[Reforecast] Re-forecast automation will be enhanced once budget baselines are established.")

    print(f"\n[Operating Rhythm] All deliverables saved to: {output_dir}")
    print("[Operating Rhythm] Done.")


if __name__ == "__main__":
    main()
