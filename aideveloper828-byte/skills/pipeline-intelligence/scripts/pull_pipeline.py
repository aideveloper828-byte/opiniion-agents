#!/usr/bin/env python3
"""
Pipeline Intelligence — Opiniion
Pulls HubSpot deal data and computes pipeline velocity, stage conversion,
deal aging, and coverage metrics. Outputs a formatted Excel workbook.

Usage:
    python3 pull_pipeline.py                            # Current open pipeline
    python3 pull_pipeline.py --closed-period 2026-Q1    # Closed deal analysis
    python3 pull_pipeline.py --quota 500000             # With quota for coverage
    python3 pull_pipeline.py --output ~/Desktop/weekly_pipeline.xlsx
"""

import argparse
import os
import sys
from datetime import datetime, timedelta

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from connectors.hubspot_connect import HubSpotClient

TEAL = "4FCBC5"
DARK_GRAY = "4C4D4E"
LIGHT_TEAL = "E0F5F4"
HEADER_GRAY = "F3F3F3"
WHITE = "FFFFFF"
STALE_RED = "FCE4D6"
GOOD_GREEN = "E2EFDA"

BRAND_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
TEAL_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
LIGHT_TEAL_FILL = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
STALE_FILL = PatternFill(start_color=STALE_RED, end_color=STALE_RED, fill_type="solid")
GOOD_FILL = PatternFill(start_color=GOOD_GREEN, end_color=GOOD_GREEN, fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_GRAY)
VALUE_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
SECTION_FILL = PatternFill(start_color=HEADER_GRAY, end_color=HEADER_GRAY, fill_type="solid")
CURRENCY_FMT = '"$"#,##0'
PCT_FMT = "0.0%"
DAYS_FMT = "#,##0"


def pull_deal_data(client: HubSpotClient) -> pd.DataFrame:
    """Pull all deals with pipeline-relevant properties."""
    props = [
        "dealname", "amount", "dealstage", "pipeline", "closedate",
        "createdate", "hs_lastmodifieddate", "hubspot_owner_id",
        "hs_deal_stage_probability", "hs_is_closed_won", "hs_is_closed",
    ]
    print("[Pipeline] Pulling all deals from HubSpot...")
    df = client.get_all_deals(properties=props)
    if df.empty:
        return df

    df["amount"] = pd.to_numeric(df.get("amount", 0), errors="coerce").fillna(0)
    df["hs_deal_stage_probability"] = pd.to_numeric(
        df.get("hs_deal_stage_probability", 0), errors="coerce"
    ).fillna(0)

    for date_col in ["closedate", "createdate", "hs_lastmodifieddate"]:
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

    return df


def pull_stage_map(client: HubSpotClient) -> dict:
    """Build a mapping of stage_id → (label, probability, display_order) from all pipelines."""
    pipelines = client.get_pipelines("deals")
    stage_map = {}
    for p in pipelines:
        for s in p.get("stages", []):
            stage_map[s["id"]] = {
                "label": s["label"],
                "probability": float(s.get("metadata", {}).get("probability", 0)),
                "display_order": s.get("displayOrder", 0),
                "pipeline": p["label"],
            }
    return stage_map


def pull_owner_map(client: HubSpotClient) -> dict:
    """Build owner_id → name mapping."""
    owners = client.get_owners()
    if owners.empty:
        return {}
    return {
        str(row["owner_id"]): f"{row['first_name']} {row['last_name']}".strip()
        for _, row in owners.iterrows()
    }


def compute_pipeline_metrics(
    deals: pd.DataFrame,
    stage_map: dict,
    owner_map: dict,
    quota: float = 0,
) -> dict:
    """Compute all pipeline metrics."""
    today = pd.Timestamp.now()

    open_deals = deals[deals.get("hs_is_closed", "false").astype(str).str.lower() != "true"].copy()
    closed_won = deals[deals.get("hs_is_closed_won", "false").astype(str).str.lower() == "true"].copy()
    closed_lost = deals[
        (deals.get("hs_is_closed", "false").astype(str).str.lower() == "true")
        & (deals.get("hs_is_closed_won", "false").astype(str).str.lower() != "true")
    ].copy()

    total_pipeline = open_deals["amount"].sum()
    weighted_pipeline = (open_deals["amount"] * open_deals["hs_deal_stage_probability"] / 100).sum()

    won_count = len(closed_won)
    lost_count = len(closed_lost)
    win_rate = won_count / (won_count + lost_count) if (won_count + lost_count) > 0 else 0

    if not closed_won.empty and "createdate" in closed_won.columns and "closedate" in closed_won.columns:
        valid_dates = closed_won.dropna(subset=["createdate", "closedate"])
        if not valid_dates.empty:
            cycle_days = (valid_dates["closedate"] - valid_dates["createdate"]).dt.days
            avg_cycle = cycle_days.mean()
        else:
            avg_cycle = 0
    else:
        avg_cycle = 0

    avg_deal_size = closed_won["amount"].mean() if not closed_won.empty else 0
    velocity = (len(open_deals) * win_rate * avg_deal_size) / avg_cycle if avg_cycle > 0 else 0
    coverage = total_pipeline / quota if quota > 0 else 0

    # Stage funnel
    stage_funnel = []
    if not open_deals.empty and "dealstage" in open_deals.columns:
        for stage_id, group in open_deals.groupby("dealstage"):
            info = stage_map.get(stage_id, {"label": stage_id, "probability": 0, "display_order": 99})
            stage_funnel.append({
                "stage_id": stage_id,
                "stage_label": info["label"],
                "display_order": info["display_order"],
                "deal_count": len(group),
                "total_value": group["amount"].sum(),
                "weighted_value": (group["amount"] * info["probability"] / 100).sum(),
                "avg_amount": group["amount"].mean(),
            })
    stage_funnel_df = pd.DataFrame(stage_funnel).sort_values("display_order") if stage_funnel else pd.DataFrame()

    # Deal aging
    open_deals_enriched = open_deals.copy()
    if not open_deals_enriched.empty:
        open_deals_enriched["stage_label"] = open_deals_enriched["dealstage"].map(
            lambda x: stage_map.get(x, {}).get("label", x)
        )
        open_deals_enriched["owner_name"] = open_deals_enriched.get("hubspot_owner_id", "").map(
            lambda x: owner_map.get(str(x), "Unassigned")
        )
        if "hs_lastmodifieddate" in open_deals_enriched.columns:
            open_deals_enriched["days_in_stage"] = (
                today - open_deals_enriched["hs_lastmodifieddate"]
            ).dt.days.fillna(0).astype(int)
        else:
            open_deals_enriched["days_in_stage"] = 0

    # Rep performance
    rep_perf = []
    all_deals = deals.copy()
    if not all_deals.empty and "hubspot_owner_id" in all_deals.columns:
        for owner_id, group in all_deals.groupby("hubspot_owner_id"):
            owner_name = owner_map.get(str(owner_id), "Unassigned")
            rep_open = group[group.get("hs_is_closed", "false").astype(str).str.lower() != "true"]
            rep_won = group[group.get("hs_is_closed_won", "false").astype(str).str.lower() == "true"]
            rep_lost = group[
                (group.get("hs_is_closed", "false").astype(str).str.lower() == "true")
                & (group.get("hs_is_closed_won", "false").astype(str).str.lower() != "true")
            ]
            rep_wr = len(rep_won) / (len(rep_won) + len(rep_lost)) if (len(rep_won) + len(rep_lost)) > 0 else 0

            rep_perf.append({
                "owner_name": owner_name,
                "open_deals": len(rep_open),
                "pipeline_value": rep_open["amount"].sum(),
                "closed_won": len(rep_won),
                "closed_lost": len(rep_lost),
                "win_rate": rep_wr,
                "total_won_value": rep_won["amount"].sum(),
            })
    rep_perf_df = pd.DataFrame(rep_perf).sort_values("pipeline_value", ascending=False) if rep_perf else pd.DataFrame()

    # Forecast categories
    commit = open_deals[open_deals["hs_deal_stage_probability"] >= 80]["amount"].sum() if not open_deals.empty else 0
    best_case_mask = (open_deals["hs_deal_stage_probability"] >= 50) & (open_deals["hs_deal_stage_probability"] < 80)
    best_case = open_deals[best_case_mask]["amount"].sum() if not open_deals.empty else 0
    pipeline_cat = open_deals[open_deals["hs_deal_stage_probability"] < 50]["amount"].sum() if not open_deals.empty else 0

    return {
        "total_pipeline": total_pipeline,
        "weighted_pipeline": weighted_pipeline,
        "open_deal_count": len(open_deals),
        "velocity": velocity,
        "win_rate": win_rate,
        "avg_cycle_days": avg_cycle,
        "avg_deal_size": avg_deal_size,
        "coverage": coverage,
        "quota": quota,
        "commit": commit,
        "best_case": best_case,
        "pipeline_category": pipeline_cat,
        "closed_won_value": closed_won["amount"].sum() if not closed_won.empty else 0,
        "stage_funnel": stage_funnel_df,
        "open_deals": open_deals_enriched,
        "rep_performance": rep_perf_df,
    }


# ── Excel Output ───────────────────────────────────────────────────

def write_summary_tab(wb: Workbook, metrics: dict):
    """Write pipeline summary dashboard."""
    ws = wb.active
    ws.title = "Pipeline Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    row = 3
    ws.cell(row=row, column=1, value="Pipeline Intelligence Report").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )
    ws.cell(row=row, column=3, value=f"As of: {datetime.today().strftime('%Y-%m-%d')}").font = VALUE_FONT
    row += 2

    kpis = [
        ("Total Pipeline", metrics["total_pipeline"], CURRENCY_FMT),
        ("Weighted Pipeline", metrics["weighted_pipeline"], CURRENCY_FMT),
        ("Open Deals", metrics["open_deal_count"], "#,##0"),
        ("Pipeline Velocity ($/day)", metrics["velocity"], CURRENCY_FMT),
        ("Win Rate", metrics["win_rate"], PCT_FMT),
        ("Avg Days to Close", metrics["avg_cycle_days"], DAYS_FMT),
        ("Avg Deal Size", metrics["avg_deal_size"], CURRENCY_FMT),
    ]
    if metrics["quota"] > 0:
        kpis.append(("Pipeline Coverage", metrics["coverage"], "0.0x"))

    ws.cell(row=row, column=1, value="Key Metrics").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=2).fill = SECTION_FILL
    row += 1

    for label, value, fmt in kpis:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.number_format = fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Forecast Categories").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=2).fill = SECTION_FILL
    row += 1

    categories = [
        ("Commit (>80%)", metrics["commit"]),
        ("Best Case (50-80%)", metrics["best_case"]),
        ("Pipeline (<50%)", metrics["pipeline_category"]),
        ("Closed Won", metrics["closed_won_value"]),
    ]
    for label, value in categories:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.number_format = CURRENCY_FMT
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A3"


def write_stage_funnel_tab(wb: Workbook, stage_funnel: pd.DataFrame):
    """Write stage funnel analysis tab."""
    ws = wb.create_sheet("Stage Funnel")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    if stage_funnel.empty:
        ws.cell(row=3, column=1, value="No open deals to analyze.").font = VALUE_FONT
        return

    headers = ["Stage", "Deal Count", "Total Value", "Weighted Value", "Avg Deal Size"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in stage_funnel.iterrows():
        r = row_idx + 4 if isinstance(row_idx, int) else stage_funnel.index.get_loc(row_idx) + 4
        ws.cell(row=r, column=1, value=row_data.get("stage_label", "")).font = VALUE_FONT
        ws.cell(row=r, column=2, value=row_data.get("deal_count", 0)).font = VALUE_FONT
        ws.cell(row=r, column=3, value=row_data.get("total_value", 0)).font = VALUE_FONT
        ws.cell(row=r, column=3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=row_data.get("weighted_value", 0)).font = VALUE_FONT
        ws.cell(row=r, column=4).number_format = CURRENCY_FMT
        ws.cell(row=r, column=5, value=row_data.get("avg_amount", 0)).font = VALUE_FONT
        ws.cell(row=r, column=5).number_format = CURRENCY_FMT

    for i, w in enumerate([22, 14, 16, 18, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def write_deal_aging_tab(wb: Workbook, open_deals: pd.DataFrame):
    """Write deal aging tab."""
    ws = wb.create_sheet("Deal Aging")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    if open_deals.empty:
        ws.cell(row=3, column=1, value="No open deals.").font = VALUE_FONT
        return

    sorted_deals = open_deals.sort_values("days_in_stage", ascending=False)
    median_days = sorted_deals["days_in_stage"].median() if not sorted_deals.empty else 30

    headers = ["Deal Name", "Owner", "Stage", "Amount", "Days in Stage", "Close Date"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, (_, deal) in enumerate(sorted_deals.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=deal.get("dealname", "")).font = VALUE_FONT
        ws.cell(row=row_idx, column=2, value=deal.get("owner_name", "Unassigned")).font = VALUE_FONT
        ws.cell(row=row_idx, column=3, value=deal.get("stage_label", "")).font = VALUE_FONT
        ws.cell(row=row_idx, column=4, value=deal.get("amount", 0)).font = VALUE_FONT
        ws.cell(row=row_idx, column=4).number_format = CURRENCY_FMT
        days_cell = ws.cell(row=row_idx, column=5, value=deal.get("days_in_stage", 0))
        days_cell.font = VALUE_FONT
        if deal.get("days_in_stage", 0) > 2 * median_days:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=STALE_RED, end_color=STALE_RED, fill_type="solid"
                )
        close_date = deal.get("closedate")
        ws.cell(row=row_idx, column=6, value=close_date.strftime("%Y-%m-%d") if pd.notna(close_date) else "").font = VALUE_FONT

    for i, w in enumerate([30, 18, 20, 14, 16, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def write_rep_performance_tab(wb: Workbook, rep_perf: pd.DataFrame):
    """Write rep performance tab."""
    ws = wb.create_sheet("Rep Performance")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    if rep_perf.empty:
        ws.cell(row=3, column=1, value="No rep data available.").font = VALUE_FONT
        return

    headers = ["Rep", "Open Deals", "Pipeline Value", "Closed Won", "Won Value", "Win Rate"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, (_, rep) in enumerate(rep_perf.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=rep.get("owner_name", "")).font = VALUE_FONT
        ws.cell(row=row_idx, column=2, value=rep.get("open_deals", 0)).font = VALUE_FONT
        ws.cell(row=row_idx, column=3, value=rep.get("pipeline_value", 0)).font = VALUE_FONT
        ws.cell(row=row_idx, column=3).number_format = CURRENCY_FMT
        ws.cell(row=row_idx, column=4, value=rep.get("closed_won", 0)).font = VALUE_FONT
        ws.cell(row=row_idx, column=5, value=rep.get("total_won_value", 0)).font = VALUE_FONT
        ws.cell(row=row_idx, column=5).number_format = CURRENCY_FMT
        ws.cell(row=row_idx, column=6, value=rep.get("win_rate", 0)).font = VALUE_FONT
        ws.cell(row=row_idx, column=6).number_format = PCT_FMT

    for i, w in enumerate([22, 14, 16, 14, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_workbook(metrics: dict, output_path: str):
    """Assemble and save the pipeline report workbook."""
    wb = Workbook()
    write_summary_tab(wb, metrics)
    write_stage_funnel_tab(wb, metrics["stage_funnel"])
    write_deal_aging_tab(wb, metrics["open_deals"])
    write_rep_performance_tab(wb, metrics["rep_performance"])
    wb.save(output_path)
    print(f"[Pipeline] Workbook saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Opiniion Pipeline Intelligence")
    parser.add_argument("--quota", type=float, default=0, help="Quota amount for coverage calculation")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    output_path = args.output or os.path.expanduser(
        f"~/Desktop/Opiniion_Pipeline_Report_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    )

    client = HubSpotClient()
    print("[Pipeline] Testing HubSpot connection...")
    if not client.test_connection():
        print("[Pipeline] Connection failed. Check .env HUBSPOT_ACCESS_TOKEN.")
        sys.exit(1)

    deals = pull_deal_data(client)
    if deals.empty:
        print("[Pipeline] WARNING: No deals returned. Building empty template.")

    stage_map = pull_stage_map(client)
    owner_map = pull_owner_map(client)

    print("[Pipeline] Computing pipeline metrics...")
    metrics = compute_pipeline_metrics(deals, stage_map, owner_map, quota=args.quota)

    print("[Pipeline] Building Excel workbook...")
    build_workbook(metrics, output_path)
    print("[Pipeline] Done.")


if __name__ == "__main__":
    main()
