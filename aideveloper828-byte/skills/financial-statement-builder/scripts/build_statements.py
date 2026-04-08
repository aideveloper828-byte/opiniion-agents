#!/usr/bin/env python3
"""
Financial Statement Builder — Opiniion
Pulls NetSuite GL actuals and generates P&L, Balance Sheet, and Cash Flow Statement.

Usage:
    python3 build_statements.py                                # YTD actuals
    python3 build_statements.py --start 2025-01 --end 2026-12 # Custom range
    python3 build_statements.py --include-forecast             # With forecast columns
    python3 build_statements.py --output ~/Desktop/Opiniion_FS.xlsx
"""

import argparse
import os
import sys
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from connectors.netsuite_connect import NetSuiteClient

# ── Opiniion brand colors & styles ─────────────────────────────────
TEAL = "4FCBC5"
DARK_GRAY = "4C4D4E"
LIGHT_TEAL = "E0F5F4"
HEADER_GRAY = "F3F3F3"
PURPLE = "7030A0"
WHITE = "FFFFFF"

BRAND_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
ACTUALS_HEADER_FILL = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
FORECAST_HEADER_FILL = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
SECTION_FILL = PatternFill(start_color=HEADER_GRAY, end_color=HEADER_GRAY, fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
VALUE_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)
FORECAST_FONT = Font(name="Calibri", size=10, color=PURPLE)
CURRENCY_FMT = '"$"#,##0'
PCT_FMT = "0.0%"
DOUBLE_BORDER = Border(bottom=Side(style="double", color=DARK_GRAY))

# ── Account mapping configuration ─────────────────────────────────
# NOTE: These are PLACEHOLDER ranges. Populate with Opiniion's actual
# account numbers after running System Discovery.

ACCOUNT_MAP = {
    "Revenue": {
        "Subscription Revenue": (4000, 4099),
        "Professional Services": (4100, 4199),
        "Other Revenue": (4200, 4299),
    },
    "Cost of Revenue": {
        "Hosting & Infrastructure": (5000, 5049),
        "Support Personnel": (5050, 5099),
        "Other COGS": (5100, 5199),
    },
    "Sales & Marketing": {"S&M Expenses": (6000, 6299)},
    "Research & Development": {"R&D Expenses": (6300, 6599)},
    "General & Administrative": {"G&A Expenses": (6600, 6999)},
    "Other Income/Expense": {
        "Interest Income": (7000, 7049),
        "Interest Expense": (7050, 7099),
        "Other": (7100, 7199),
    },
    "Income Tax": {"Tax Expense": (8000, 8099)},
}

EBITDA_PATTERNS = {
    "Depreciation": ["depreciation", "depr"],
    "Amortization": ["amortization", "amort"],
    "Stock-Based Compensation": ["stock comp", "sbc", "stock-based"],
}

BS_ACCOUNT_MAP = {
    "Current Assets": {
        "Cash & Equivalents": (1000, 1099),
        "Accounts Receivable": (1100, 1199),
        "Prepaid Expenses": (1200, 1299),
        "Other Current Assets": (1300, 1499),
    },
    "Fixed Assets": {
        "Property & Equipment": (1500, 1599),
        "Accumulated Depreciation": (1600, 1699),
    },
    "Long-Term Assets": {
        "Intangible Assets": (1700, 1799),
        "Other LT Assets": (1800, 1899),
    },
    "Current Liabilities": {
        "Accounts Payable": (2000, 2099),
        "Accrued Liabilities": (2100, 2199),
        "Deferred Revenue": (2200, 2299),
        "Current Debt": (2300, 2399),
    },
    "Long-Term Liabilities": {
        "Long-Term Debt": (2500, 2599),
        "Other LT Liabilities": (2600, 2699),
    },
    "Equity": {
        "Common Stock": (3000, 3049),
        "APIC": (3050, 3099),
        "Retained Earnings": (3100, 3149),
        "Treasury Stock": (3200, 3249),
    },
}


def classify_account(acct_number: str, acct_name: str) -> tuple[str, str]:
    """Classify a NetSuite account into financial statement section and line item."""
    try:
        num = int(re.sub(r"[^0-9]", "", str(acct_number))[:4])
    except (ValueError, IndexError):
        return "Unmapped", acct_name

    for section, lines in ACCOUNT_MAP.items():
        for line_item, (low, high) in lines.items():
            if low <= num <= high:
                return section, line_item

    for section, lines in BS_ACCOUNT_MAP.items():
        for line_item, (low, high) in lines.items():
            if low <= num <= high:
                return section, line_item

    return "Unmapped", acct_name


def pull_gl_data(client: NetSuiteClient, start_date: str, end_date: str) -> pd.DataFrame:
    """Pull GL trial balance data from NetSuite by account and period."""
    sql = f"""
    SELECT
        a.acctnumber AS account_number,
        a.acctname AS account_name,
        a.accttype AS account_type,
        ap.periodname AS period_name,
        ap.startdate AS period_start,
        SUM(tal.amount) AS amount
    FROM transactionaccountingline tal
    JOIN transaction t ON tal.transaction = t.id
    JOIN account a ON tal.account = a.id
    JOIN accountingperiod ap ON tal.postingperiod = ap.id
    WHERE ap.startdate >= '{start_date}'
      AND ap.startdate < '{end_date}'
      AND ap.isadjust = 'F'
      AND ap.isquarter = 'F'
      AND ap.isyear = 'F'
    GROUP BY a.acctnumber, a.acctname, a.accttype, ap.periodname, ap.startdate
    ORDER BY ap.startdate, a.acctnumber
    """
    print("[FS Builder] Pulling GL data from NetSuite...")
    df = client.query(sql)
    if df.empty:
        return df

    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
    classified = df.apply(
        lambda row: classify_account(row["account_number"], row["account_name"]),
        axis=1, result_type="expand"
    )
    df["section"] = classified[0]
    df["line_item"] = classified[1]
    return df


def build_pl_data(gl: pd.DataFrame) -> pd.DataFrame:
    """Aggregate GL data into P&L line items by month."""
    pl_sections = list(ACCOUNT_MAP.keys())
    pl_data = gl[gl["section"].isin(pl_sections)].copy()

    if pl_data.empty:
        return pd.DataFrame()

    # Revenue accounts have credit balances (negative) — flip sign
    revenue_mask = pl_data["section"] == "Revenue"
    pl_data.loc[revenue_mask, "amount"] = -pl_data.loc[revenue_mask, "amount"]

    summary = pl_data.pivot_table(
        index=["section", "line_item"],
        columns="period_start",
        values="amount",
        aggfunc="sum",
        fill_value=0,
    )
    return summary


def build_bs_data(gl: pd.DataFrame) -> pd.DataFrame:
    """Aggregate GL data into Balance Sheet line items (latest period only)."""
    bs_sections = list(BS_ACCOUNT_MAP.keys())
    bs_data = gl[gl["section"].isin(bs_sections)].copy()

    if bs_data.empty:
        return pd.DataFrame()

    summary = bs_data.pivot_table(
        index=["section", "line_item"],
        columns="period_start",
        values="amount",
        aggfunc="sum",
        fill_value=0,
    )
    return summary


# ── Excel Output ───────────────────────────────────────────────────

def write_pl_tab(wb: Workbook, pl: pd.DataFrame, months: list):
    """Write the Income Statement tab."""
    ws = wb.active
    ws.title = "Income Statement"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    if pl.empty:
        ws.cell(row=3, column=1, value="No GL data available. Check SuiteQL queries and account mapping.").font = VALUE_FONT
        return

    header_row = 3
    ws.cell(row=header_row, column=1, value="").font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = ACTUALS_HEADER_FILL
    for col_idx, month in enumerate(months, 2):
        month_label = pd.Timestamp(month).strftime("%b-%y")
        cell = ws.cell(row=header_row, column=col_idx, value=month_label)
        cell.font = HEADER_FONT
        cell.fill = ACTUALS_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ytd_col = len(months) + 2
    cell = ws.cell(row=header_row, column=ytd_col, value="YTD Total")
    cell.font = HEADER_FONT
    cell.fill = ACTUALS_HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

    row = 4
    pl_order = ["Revenue", "Cost of Revenue", "Sales & Marketing",
                "Research & Development", "General & Administrative",
                "Other Income/Expense", "Income Tax"]

    for section in pl_order:
        if section not in pl.index.get_level_values(0):
            continue

        ws.cell(row=row, column=1, value=section).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        for c in range(2, ytd_col + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        section_data = pl.loc[section] if section in pl.index.get_level_values(0) else pd.DataFrame()
        if isinstance(section_data, pd.Series):
            section_data = section_data.to_frame().T

        for line_item in section_data.index:
            label = line_item if isinstance(line_item, str) else line_item[0] if isinstance(line_item, tuple) else str(line_item)
            ws.cell(row=row, column=1, value=f"  {label}").font = VALUE_FONT
            for col_idx, month in enumerate(months, 2):
                val = section_data.loc[line_item, month] if month in section_data.columns else 0
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = VALUE_FONT
                cell.number_format = CURRENCY_FMT

            # YTD formula
            first_col = get_column_letter(2)
            last_col = get_column_letter(len(months) + 1)
            ytd_cell = ws.cell(row=row, column=ytd_col)
            ytd_cell.value = f"=SUM({first_col}{row}:{last_col}{row})"
            ytd_cell.font = VALUE_FONT
            ytd_cell.number_format = CURRENCY_FMT
            row += 1

        # Section subtotal
        section_start_row = row - len(section_data.index) if hasattr(section_data, 'index') else row
        ws.cell(row=row, column=1, value=f"Total {section}").font = Font(
            name="Calibri", size=10, bold=True, color=DARK_GRAY
        )
        for c in range(1, ytd_col + 1):
            ws.cell(row=row, column=c).fill = SUBTOTAL_FILL
        for col_idx in range(2, ytd_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=row, column=col_idx).value = f"=SUM({col_letter}{section_start_row}:{col_letter}{row - 1})"
            ws.cell(row=row, column=col_idx).number_format = CURRENCY_FMT
            ws.cell(row=row, column=col_idx).font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        row += 2

    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, ytd_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.freeze_panes = "B4"


def write_bs_tab(wb: Workbook, bs: pd.DataFrame, months: list):
    """Write the Balance Sheet tab."""
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    if bs.empty:
        ws.cell(row=3, column=1, value="No Balance Sheet data available.").font = VALUE_FONT
        return

    header_row = 3
    ws.cell(row=header_row, column=1, value="").fill = ACTUALS_HEADER_FILL
    for col_idx, month in enumerate(months, 2):
        cell = ws.cell(row=header_row, column=col_idx, value=pd.Timestamp(month).strftime("%b-%y"))
        cell.font = HEADER_FONT
        cell.fill = ACTUALS_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 4
    bs_order = ["Current Assets", "Fixed Assets", "Long-Term Assets",
                "Current Liabilities", "Long-Term Liabilities", "Equity"]

    for section in bs_order:
        if section not in bs.index.get_level_values(0):
            continue
        ws.cell(row=row, column=1, value=section).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        row += 1

        section_data = bs.loc[section]
        if isinstance(section_data, pd.Series):
            section_data = section_data.to_frame().T

        for line_item in section_data.index:
            label = line_item if isinstance(line_item, str) else str(line_item)
            ws.cell(row=row, column=1, value=f"  {label}").font = VALUE_FONT
            for col_idx, month in enumerate(months, 2):
                val = section_data.loc[line_item, month] if month in section_data.columns else 0
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = VALUE_FONT
                cell.number_format = CURRENCY_FMT
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(months) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.freeze_panes = "B4"


def write_cf_tab(wb: Workbook):
    """Write the Cash Flow Statement tab (placeholder structure)."""
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    ws.cell(row=3, column=1, value="Cash Flow Statement").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )
    ws.cell(row=5, column=1, value="Cash Flow Statement will be derived from P&L and Balance Sheet").font = VALUE_FONT
    ws.cell(row=6, column=1, value="changes once 2+ months of Balance Sheet data are available.").font = VALUE_FONT

    sections = [
        (8, "Operating Activities"),
        (9, "  Net Income"),
        (10, "  + Depreciation & Amortization"),
        (11, "  + Stock-Based Compensation"),
        (12, "  +/- Working Capital Changes"),
        (13, "Net Cash from Operations"),
        (15, "Investing Activities"),
        (16, "  Capital Expenditures"),
        (17, "  Capitalized Software"),
        (18, "Net Cash from Investing"),
        (20, "Financing Activities"),
        (21, "  Debt Proceeds / (Repayments)"),
        (22, "  Equity Issuance / (Repurchase)"),
        (23, "Net Cash from Financing"),
        (25, "Net Change in Cash"),
    ]

    for row_num, label in sections:
        if label.startswith("Net"):
            ws.cell(row=row_num, column=1, value=label).font = Font(
                name="Calibri", size=10, bold=True, color=DARK_GRAY
            )
            ws.cell(row=row_num, column=1).border = DOUBLE_BORDER
        elif not label.startswith("  "):
            ws.cell(row=row_num, column=1, value=label).font = SECTION_FONT
            ws.cell(row=row_num, column=1).fill = SECTION_FILL
        else:
            ws.cell(row=row_num, column=1, value=label).font = VALUE_FONT

    ws.column_dimensions["A"].width = 35


def write_assumptions_tab(wb: Workbook):
    """Write the editable Assumptions tab."""
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "opiniion"
    ws["A1"].font = BRAND_FONT

    ws.cell(row=3, column=1, value="Forecast Assumptions").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )

    INPUT_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    INPUT_FONT = Font(name="Calibri", size=9, bold=True, color="0000FF")

    assumptions = [
        ("Revenue Growth Rate (Annual)", 0.20, PCT_FMT),
        ("Gross Margin %", 0.70, PCT_FMT),
        ("S&M as % of Revenue", 0.40, PCT_FMT),
        ("R&D as % of Revenue", 0.20, PCT_FMT),
        ("G&A as % of Revenue", 0.15, PCT_FMT),
        ("Tax Rate", 0.25, PCT_FMT),
    ]

    row = 5
    ws.cell(row=row, column=1, value="Assumption").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=2, value="Value").font = SECTION_FONT
    ws.cell(row=row, column=2).fill = SECTION_FILL
    row += 1

    for label, default_val, fmt in assumptions:
        ws.cell(row=row, column=1, value=label).font = VALUE_FONT
        cell = ws.cell(row=row, column=2, value=default_val)
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL
        cell.number_format = fmt
        row += 1

    ws.cell(row=row + 1, column=1, value="Green cells are editable inputs.").font = Font(
        name="Calibri", size=9, italic=True, color="999999"
    )
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14


def build_workbook(gl: pd.DataFrame, output_path: str, months: list):
    """Assemble and save the full financial statements workbook."""
    wb = Workbook()

    pl = build_pl_data(gl)
    bs = build_bs_data(gl)

    write_pl_tab(wb, pl, months)
    write_bs_tab(wb, bs, months)
    write_cf_tab(wb)
    write_assumptions_tab(wb)

    wb.save(output_path)
    print(f"[FS Builder] Workbook saved to: {output_path}")

    unmapped = gl[gl["section"] == "Unmapped"]
    if not unmapped.empty:
        print(f"[FS Builder] WARNING: {len(unmapped)} rows in {unmapped['account_number'].nunique()} accounts are UNMAPPED.")
        print("[FS Builder] Update account-mapping.md and re-run.")


def main():
    parser = argparse.ArgumentParser(description="Opiniion Financial Statement Builder")
    parser.add_argument("--start", default=f"{datetime.today().year}-01", help="Start period (YYYY-MM)")
    parser.add_argument("--end", default=datetime.today().strftime("%Y-%m"), help="End period (YYYY-MM)")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    start_date = f"{args.start}-01"
    from dateutil.relativedelta import relativedelta
    end_dt = datetime.strptime(f"{args.end}-01", "%Y-%m-%d") + relativedelta(months=1)
    end_date = end_dt.strftime("%Y-%m-%d")

    output_path = args.output or os.path.expanduser(
        f"~/Desktop/Opiniion_Financial_Statements_{datetime.today().strftime('%Y-%m')}.xlsx"
    )

    client = NetSuiteClient()
    print("[FS Builder] Testing NetSuite connection...")
    if not client.test_connection():
        print("[FS Builder] Connection failed. Check .env credentials.")
        sys.exit(1)

    gl = pull_gl_data(client, start_date, end_date)
    if gl.empty:
        print("[FS Builder] WARNING: No GL data returned. Building empty template.")

    months = sorted(gl["period_start"].unique()) if not gl.empty else []

    print("[FS Builder] Building workbook...")
    build_workbook(gl, output_path, months)
    print("[FS Builder] Done.")


if __name__ == "__main__":
    main()
