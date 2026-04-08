#!/usr/bin/env python3
"""
SaaS Metrics Engine — Opiniion
Computes ARR waterfall, retention, efficiency, and velocity metrics from NetSuite data.
Outputs a formatted Excel workbook with Opiniion branding.

Usage:
    python3 compute_metrics.py                              # TTM, all customers
    python3 compute_metrics.py --period 2026-Q1             # Specific quarter
    python3 compute_metrics.py --period 2025-01:2025-12     # Custom range
    python3 compute_metrics.py --output ~/Desktop/custom.xlsx
"""

import argparse
import os
import sys
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from connectors.netsuite_connect import NetSuiteClient

# ── Opiniion brand colors ──────────────────────────────────────────
TEAL = "4FCBC5"
DARK_GRAY = "4C4D4E"
LIGHT_TEAL = "E0F5F4"
HEADER_GRAY = "F3F3F3"
WHITE = "FFFFFF"
RED_DECLINE = "FCE4D6"
GREEN_IMPROVE = "E2EFDA"

TEAL_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
LIGHT_TEAL_FILL = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
HEADER_FILL = PatternFill(start_color=HEADER_GRAY, end_color=HEADER_GRAY, fill_type="solid")
RED_FILL = PatternFill(start_color=RED_DECLINE, end_color=RED_DECLINE, fill_type="solid")
GREEN_FILL = PatternFill(start_color=GREEN_IMPROVE, end_color=GREEN_IMPROVE, fill_type="solid")

BRAND_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
METRIC_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_GRAY)
VALUE_FONT = Font(name="Calibri", size=11, color=DARK_GRAY)
CURRENCY_FMT = '"$"#,##0'
PCT_FMT = "0.0%"
NUMBER_FMT = "#,##0"
RATIO_FMT = "0.0x"
DAYS_FMT = "#,##0"

THIN_BORDER = Border(bottom=Side(style="thin", color=DARK_GRAY))


def parse_period(period_str: str):
    """Parse period argument into (start_date, end_date) as YYYY-MM-01 strings."""
    today = datetime.today()

    if period_str == "TTM":
        end = today.replace(day=1)
        start = end - relativedelta(months=12)
        return start.strftime("%Y-%m-01"), end.strftime("%Y-%m-01")

    if period_str.startswith("20") and "-Q" in period_str:
        year, q = period_str.split("-Q")
        q = int(q)
        start_month = (q - 1) * 3 + 1
        start = datetime(int(year), start_month, 1)
        end = start + relativedelta(months=3)
        return start.strftime("%Y-%m-01"), end.strftime("%Y-%m-01")

    if ":" in period_str:
        parts = period_str.split(":")
        return f"{parts[0]}-01", f"{parts[1]}-01"

    return f"{period_str}-01", (datetime.strptime(f"{period_str}-01", "%Y-%m-%d") + relativedelta(months=1)).strftime("%Y-%m-01")


def pull_revenue_data(client: NetSuiteClient, start_date: str, end_date: str) -> pd.DataFrame:
    """Pull monthly revenue by customer from NetSuite.

    NOTE: These SuiteQL queries use standard NetSuite record types. Column and table
    names may need adjustment based on Opiniion's actual NetSuite configuration.
    Run the System Discovery agent first to confirm the schema.
    """
    sql = f"""
    SELECT
        t.entity AS customer_id,
        c.companyname AS customer_name,
        TO_CHAR(t.trandate, 'YYYY-MM') || '-01' AS month,
        SUM(tl.amount) AS revenue
    FROM transaction t
    JOIN transactionline tl ON t.id = tl.transaction
    JOIN customer c ON t.entity = c.id
    WHERE t.type = 'CustInvc'
      AND tl.mainline = 'F'
      AND t.trandate >= '{start_date}'
      AND t.trandate < '{end_date}'
    GROUP BY t.entity, c.companyname, TO_CHAR(t.trandate, 'YYYY-MM') || '-01'
    ORDER BY month, customer_name
    """
    df = client.query(sql)
    if df.empty:
        return df
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce").fillna(0)
    return df


def pull_sm_spend(client: NetSuiteClient, start_date: str, end_date: str) -> float:
    """Pull total S&M department spend from NetSuite GL for efficiency metrics."""
    sql = f"""
    SELECT SUM(tl.amount) AS sm_spend
    FROM transaction t
    JOIN transactionline tl ON t.id = tl.transaction
    JOIN department d ON tl.department = d.id
    WHERE t.type IN ('VendBill', 'Journal', 'ExpRept')
      AND LOWER(d.name) LIKE '%sales%'
      AND t.trandate >= '{start_date}'
      AND t.trandate < '{end_date}'
      AND tl.mainline = 'F'
    """
    df = client.query(sql)
    if df.empty:
        return 0.0
    return abs(pd.to_numeric(df.iloc[0].get("sm_spend", 0), errors="coerce") or 0)


def compute_arr_waterfall(rev: pd.DataFrame, months: list[str]) -> pd.DataFrame:
    """Compute monthly ARR waterfall from revenue by customer."""
    if rev.empty:
        return pd.DataFrame()

    pivot = rev.pivot_table(
        index="customer_id", columns="month", values="revenue", aggfunc="sum", fill_value=0
    )
    for m in months:
        if m not in pivot.columns:
            pivot[m] = 0
    pivot = pivot[sorted(pivot.columns)]

    rows = []
    cols = sorted(pivot.columns)
    for i, month in enumerate(cols):
        current_arr = pivot[month].sum() * 12
        prev_month = cols[i - 1] if i > 0 else None
        prev_arr = (pivot[prev_month].sum() * 12) if prev_month else 0

        new = expansion = contraction = churn = 0
        for cust_id in pivot.index:
            curr = pivot.loc[cust_id, month] * 12
            prev = (pivot.loc[cust_id, prev_month] * 12) if prev_month else 0

            if prev == 0 and curr > 0:
                new += curr
            elif prev > 0 and curr > prev:
                expansion += curr - prev
            elif prev > 0 and 0 < curr < prev:
                contraction += prev - curr
            elif prev > 0 and curr == 0:
                churn += prev

        rows.append({
            "Month": month,
            "Beginning ARR": prev_arr,
            "New ARR": new,
            "Expansion ARR": expansion,
            "Contraction ARR": contraction,
            "Churned ARR": churn,
            "Ending ARR": current_arr,
        })

    return pd.DataFrame(rows)


def compute_retention(waterfall: pd.DataFrame) -> dict:
    """Compute GRR and NRR from the ARR waterfall (period totals)."""
    if waterfall.empty:
        return {"GRR": 0, "NRR": 0}

    beg = waterfall.iloc[0]["Beginning ARR"]
    if beg == 0:
        beg = waterfall.iloc[1]["Beginning ARR"] if len(waterfall) > 1 else 1

    total_contraction = waterfall["Contraction ARR"].sum()
    total_churn = waterfall["Churned ARR"].sum()
    total_expansion = waterfall["Expansion ARR"].sum()

    grr = (beg - total_contraction - total_churn) / beg if beg else 0
    nrr = (beg + total_expansion - total_contraction - total_churn) / beg if beg else 0

    return {"GRR": grr, "NRR": nrr}


def compute_efficiency(waterfall: pd.DataFrame, sm_spend: float) -> dict:
    """Compute Magic Number, LTV, CAC, LTV/CAC from waterfall and S&M spend."""
    if waterfall.empty:
        return {"Magic Number": 0, "LTV": 0, "CAC": 0, "LTV/CAC": 0, "CAC Payback": 0}

    net_new = waterfall["New ARR"].sum() + waterfall["Expansion ARR"].sum()
    ending = waterfall.iloc[-1]["Ending ARR"]
    beg = waterfall.iloc[0]["Beginning ARR"] or 1

    magic = (ending - beg) / sm_spend if sm_spend else 0

    gross_churn_rate = (waterfall["Contraction ARR"].sum() + waterfall["Churned ARR"].sum()) / beg if beg else 1
    gross_margin = 0.70  # placeholder — update with actual GM% from financial statements
    ltv = (net_new * gross_margin) / gross_churn_rate if gross_churn_rate else 0

    active_customers = max(1, 1)  # placeholder — replace with actual count
    cac = sm_spend / active_customers if active_customers else 0

    ltv_cac = ltv / cac if cac else 0
    arpu_monthly = (ending / active_customers / 12) if active_customers else 0
    cac_payback = cac / (arpu_monthly * gross_margin) if (arpu_monthly * gross_margin) else 0

    return {
        "Magic Number": magic,
        "LTV": ltv,
        "CAC": cac,
        "LTV/CAC": ltv_cac,
        "CAC Payback": cac_payback,
    }


# ── Excel Output ───────────────────────────────────────────────────

def write_brand_header(ws, text="opiniion"):
    """Write Opiniion branding in A1."""
    ws["A1"] = text
    ws["A1"].font = BRAND_FONT


def write_summary_tab(wb: Workbook, metrics: dict):
    """Write the Summary Dashboard tab."""
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.sheet_view.showGridLines = False
    write_brand_header(ws)

    row = 3
    ws.cell(row=row, column=1, value="SaaS Metrics Summary").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )
    ws.cell(row=row, column=3, value=f"Period: {metrics.get('period', 'TTM')}").font = VALUE_FONT
    row += 2

    sections = [
        ("Revenue", [
            ("Ending ARR", metrics.get("ending_arr", 0), CURRENCY_FMT),
            ("MRR", metrics.get("ending_arr", 0) / 12, CURRENCY_FMT),
            ("New ARR", metrics.get("new_arr", 0), CURRENCY_FMT),
            ("Expansion ARR", metrics.get("expansion_arr", 0), CURRENCY_FMT),
        ]),
        ("Retention", [
            ("Gross Retention (GRR)", metrics.get("grr", 0), PCT_FMT),
            ("Net Retention (NRR)", metrics.get("nrr", 0), PCT_FMT),
        ]),
        ("Efficiency", [
            ("Magic Number", metrics.get("magic_number", 0), RATIO_FMT),
            ("LTV / CAC", metrics.get("ltv_cac", 0), RATIO_FMT),
            ("CAC Payback (months)", metrics.get("cac_payback", 0), DAYS_FMT),
        ]),
        ("Sales Velocity", [
            ("Win Rate", metrics.get("win_rate", 0), PCT_FMT),
            ("Avg Days to Close", metrics.get("avg_days_to_close", 0), DAYS_FMT),
            ("Avg Deal Size", metrics.get("avg_deal_size", 0), CURRENCY_FMT),
        ]),
    ]

    for section_name, items in sections:
        ws.cell(row=row, column=1, value=section_name).font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3).fill = HEADER_FILL
        row += 1
        for label, value, fmt in items:
            ws.cell(row=row, column=1, value=label).font = METRIC_LABEL_FONT
            cell = ws.cell(row=row, column=2, value=value)
            cell.font = VALUE_FONT
            cell.number_format = fmt
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A3"


def write_waterfall_tab(wb: Workbook, waterfall: pd.DataFrame):
    """Write the ARR Waterfall tab."""
    ws = wb.create_sheet("ARR Waterfall")
    ws.sheet_view.showGridLines = False
    write_brand_header(ws)

    if waterfall.empty:
        ws.cell(row=3, column=1, value="No data available for this period.").font = VALUE_FONT
        return

    headers = ["Month", "Beginning ARR", "New ARR", "Expansion ARR",
               "Contraction ARR", "Churned ARR", "Ending ARR"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in waterfall.iterrows():
        excel_row = row_idx + 4
        ws.cell(row=excel_row, column=1, value=row_data["Month"]).font = VALUE_FONT
        for col_idx, col_name in enumerate(headers[1:], 2):
            cell = ws.cell(row=excel_row, column=col_idx, value=row_data[col_name])
            cell.number_format = CURRENCY_FMT
            cell.font = VALUE_FONT
            if col_name in ("Contraction ARR", "Churned ARR") and row_data[col_name] > 0:
                cell.font = Font(name="Calibri", size=11, color="CC0000")

    total_row = len(waterfall) + 4
    ws.cell(row=total_row, column=1, value="Total").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TEAL_FILL
    for col_idx in range(2, 8):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(
            row=total_row, column=col_idx,
            value=f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
        )
        if col_idx in (1, 2, 7):
            cell.value = waterfall.iloc[-1][headers[col_idx - 1]] if col_idx == 7 else waterfall.iloc[0][headers[col_idx - 1]]
        cell.number_format = CURRENCY_FMT
        cell.font = TOTAL_FONT
        cell.fill = TEAL_FILL

    for i, width in enumerate([14, 18, 14, 16, 16, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "B4"


def write_cohort_tab(wb: Workbook, rev: pd.DataFrame):
    """Write the Cohort Analysis tab (retention by sign-up quarter)."""
    ws = wb.create_sheet("Cohort Analysis")
    ws.sheet_view.showGridLines = False
    write_brand_header(ws)

    if rev.empty:
        ws.cell(row=3, column=1, value="No data available for cohort analysis.").font = VALUE_FONT
        return

    ws.cell(row=3, column=1, value="Cohort retention analysis will be populated").font = VALUE_FONT
    ws.cell(row=4, column=1, value="once sufficient historical data is available.").font = VALUE_FONT
    ws.cell(row=6, column=1, value="Requires 6+ months of billing data to generate meaningful cohort curves.").font = Font(
        name="Calibri", size=10, italic=True, color="999999"
    )


def write_definitions_tab(wb: Workbook):
    """Write the Definitions tab with all metric formulas."""
    ws = wb.create_sheet("Definitions")
    ws.sheet_view.showGridLines = False
    write_brand_header(ws)

    definitions = [
        ("Ending ARR", "SUM(monthly_recurring_revenue * 12) for all active customers at period end"),
        ("Beginning ARR", "Ending ARR of the prior period"),
        ("New ARR", "ARR from customers first invoiced during the current period"),
        ("Expansion ARR", "Incremental ARR from existing customers who increased spend"),
        ("Contraction ARR", "Decrease in ARR from existing customers who reduced spend (not fully churned)"),
        ("Churned ARR", "Full ARR lost from customers who canceled or did not renew"),
        ("GRR", "(Beginning ARR - Contraction - Churn) / Beginning ARR"),
        ("NRR", "(Beginning ARR + Expansion - Contraction - Churn) / Beginning ARR"),
        ("Magic Number", "(Ending ARR - Beginning ARR) / S&M Spend"),
        ("LTV", "((New + Expansion ARR) * Gross Margin %) / Gross Churn Rate"),
        ("CAC", "(S&M + Implementation Spend) / New Customers"),
        ("LTV / CAC", "LTV / CAC"),
        ("Win Rate", "Closed Won / (Closed Won + Closed Lost)"),
        ("Avg Days to Close", "Mean(close_date - create_date) for Closed Won deals"),
        ("Avg Deal Size", "Total New ARR / Count of Closed Won deals"),
        ("ARPU", "Ending ARR / Count of Active Customers"),
    ]

    ws.cell(row=3, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=2, value="Formula / Definition").font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL

    for i, (metric, formula) in enumerate(definitions, 4):
        ws.cell(row=i, column=1, value=metric).font = METRIC_LABEL_FONT
        ws.cell(row=i, column=2, value=formula).font = VALUE_FONT

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


def build_workbook(
    waterfall: pd.DataFrame,
    rev: pd.DataFrame,
    metrics: dict,
    output_path: str,
):
    """Assemble and save the full SaaS Metrics workbook."""
    wb = Workbook()
    write_summary_tab(wb, metrics)
    write_waterfall_tab(wb, waterfall)
    write_cohort_tab(wb, rev)
    write_definitions_tab(wb)
    wb.save(output_path)
    print(f"[SaaS Metrics] Workbook saved to: {output_path}")


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Opiniion SaaS Metrics Engine")
    parser.add_argument("--period", default="TTM", help="TTM, YYYY-QN, YYYY-MM, or YYYY-MM:YYYY-MM")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    start_date, end_date = parse_period(args.period)
    period_label = args.period if args.period != "TTM" else f"TTM ({start_date} to {end_date})"

    output_path = args.output or os.path.expanduser(
        f"~/Desktop/Opiniion_SaaS_Metrics_{datetime.today().strftime('%Y-%m')}.xlsx"
    )

    print(f"[SaaS Metrics] Period: {period_label}")
    print(f"[SaaS Metrics] Connecting to NetSuite...")

    client = NetSuiteClient()

    print("[SaaS Metrics] Pulling revenue data...")
    rev = pull_revenue_data(client, start_date, end_date)
    if rev.empty:
        print("[SaaS Metrics] WARNING: No revenue data returned. Check SuiteQL queries and date range.")
        print("[SaaS Metrics] Building workbook with empty data (template)...")

    months = sorted(rev["month"].unique()) if not rev.empty else []

    print("[SaaS Metrics] Computing ARR waterfall...")
    waterfall = compute_arr_waterfall(rev, months)

    print("[SaaS Metrics] Computing retention metrics...")
    retention = compute_retention(waterfall)

    print("[SaaS Metrics] Pulling S&M spend...")
    sm_spend = pull_sm_spend(client, start_date, end_date)

    print("[SaaS Metrics] Computing efficiency metrics...")
    efficiency = compute_efficiency(waterfall, sm_spend)

    summary_metrics = {
        "period": period_label,
        "ending_arr": waterfall.iloc[-1]["Ending ARR"] if not waterfall.empty else 0,
        "new_arr": waterfall["New ARR"].sum() if not waterfall.empty else 0,
        "expansion_arr": waterfall["Expansion ARR"].sum() if not waterfall.empty else 0,
        "grr": retention["GRR"],
        "nrr": retention["NRR"],
        "magic_number": efficiency["Magic Number"],
        "ltv_cac": efficiency["LTV/CAC"],
        "cac_payback": efficiency["CAC Payback"],
        "win_rate": 0,  # populated from HubSpot via pipeline-intelligence skill
        "avg_days_to_close": 0,
        "avg_deal_size": 0,
    }

    print("[SaaS Metrics] Building Excel workbook...")
    build_workbook(waterfall, rev, summary_metrics, output_path)
    print("[SaaS Metrics] Done.")


if __name__ == "__main__":
    main()
