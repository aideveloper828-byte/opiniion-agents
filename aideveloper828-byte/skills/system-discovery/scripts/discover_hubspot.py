#!/usr/bin/env python3
"""
HubSpot System Discovery — Opiniion
Catalogs pipelines, stages, deal/contact/company properties, owners, and custom objects.
Outputs a formatted Data Dictionary Excel workbook.

Usage:
    python3 discover_hubspot.py
    python3 discover_hubspot.py --output ~/Desktop/HS_Discovery.xlsx
"""

import argparse
import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from connectors.hubspot_connect import HubSpotClient

TEAL = "4FCBC5"
DARK_GRAY = "4C4D4E"
BRAND_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
VALUE_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)


def discover_pipelines(client: HubSpotClient) -> pd.DataFrame:
    """Pull all deal pipelines with their stages."""
    print("[Discovery] Pulling deal pipelines and stages...")
    pipelines = client.get_pipelines("deals")
    rows = []
    for p in pipelines:
        for s in p.get("stages", []):
            rows.append({
                "pipeline_id": p["id"],
                "pipeline_label": p["label"],
                "stage_id": s["id"],
                "stage_label": s["label"],
                "display_order": s.get("displayOrder", 0),
                "probability": s.get("metadata", {}).get("probability", ""),
            })
    return pd.DataFrame(rows)


def discover_properties(client: HubSpotClient, object_type: str) -> pd.DataFrame:
    """Pull all properties for a given object type."""
    print(f"[Discovery] Pulling {object_type} properties...")
    df = client.get_properties(object_type)
    if not df.empty:
        df["options"] = df["options"].apply(lambda x: "; ".join(x) if isinstance(x, list) else str(x))
    return df


def discover_owners(client: HubSpotClient) -> pd.DataFrame:
    """Pull all HubSpot owners (sales reps)."""
    print("[Discovery] Pulling owners...")
    return client.get_owners()


def discover_record_counts(client: HubSpotClient) -> pd.DataFrame:
    """Get approximate record counts for each object type."""
    print("[Discovery] Counting records by object type...")
    counts = []
    for obj_type in ["deals", "contacts", "companies"]:
        try:
            import requests
            resp = client._request("GET", f"/crm/v3/objects/{obj_type}", params={"limit": 1})
            total = resp.json().get("total", "unknown")
            counts.append({"object_type": obj_type, "record_count": total})
        except Exception:
            counts.append({"object_type": obj_type, "record_count": "error"})
    return pd.DataFrame(counts)


def write_discovery_workbook(
    pipelines: pd.DataFrame,
    deal_props: pd.DataFrame,
    contact_props: pd.DataFrame,
    company_props: pd.DataFrame,
    owners: pd.DataFrame,
    record_counts: pd.DataFrame,
    output_path: str,
):
    """Write the HubSpot discovery workbook."""
    wb = Workbook()

    datasets = [
        ("Deal Pipelines", pipelines),
        ("Deal Properties", deal_props),
        ("Contact Properties", contact_props),
        ("Company Properties", company_props),
        ("Owners", owners),
        ("Record Counts", record_counts),
    ]

    for idx, (tab_name, df) in enumerate(datasets):
        if idx == 0:
            ws = wb.active
            ws.title = tab_name
        else:
            ws = wb.create_sheet(tab_name)

        ws.sheet_view.showGridLines = False
        ws["A1"] = "opiniion"
        ws["A1"].font = BRAND_FONT

        if df.empty:
            ws.cell(row=3, column=1, value="No data returned. Check API token permissions.").font = VALUE_FONT
            continue

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 4, column=col_idx, value=str(value) if value is not None else "")
                cell.font = VALUE_FONT

        for col_idx in range(1, len(df.columns) + 1):
            max_len = max(
                len(str(df.columns[col_idx - 1])),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if not df.empty else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A4"

    summary = wb.create_sheet("Summary", 0)
    summary.sheet_view.showGridLines = False
    summary["A1"] = "opiniion"
    summary["A1"].font = BRAND_FONT
    summary.cell(row=3, column=1, value="HubSpot Data Discovery Summary").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )
    summary.cell(row=4, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d %H:%M')}").font = VALUE_FONT

    row = 6
    for tab_name, df in datasets:
        summary.cell(row=row, column=1, value=tab_name).font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        summary.cell(row=row, column=2, value=f"{len(df)} records").font = VALUE_FONT
        row += 1

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 15

    wb.save(output_path)
    print(f"[Discovery] HubSpot workbook saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="HubSpot System Discovery")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    output_path = args.output or os.path.expanduser(
        f"~/Desktop/Opiniion_HubSpot_Discovery_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    )

    client = HubSpotClient()
    print("[Discovery] Testing HubSpot connection...")
    if not client.test_connection():
        print("[Discovery] Connection failed. Check .env HUBSPOT_ACCESS_TOKEN.")
        sys.exit(1)

    pipelines = discover_pipelines(client)
    deal_props = discover_properties(client, "deals")
    contact_props = discover_properties(client, "contacts")
    company_props = discover_properties(client, "companies")
    owners = discover_owners(client)
    record_counts = discover_record_counts(client)

    write_discovery_workbook(
        pipelines, deal_props, contact_props, company_props, owners, record_counts, output_path
    )
    print("[Discovery] Done.")


if __name__ == "__main__":
    main()
