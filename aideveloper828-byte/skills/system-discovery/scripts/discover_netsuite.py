#!/usr/bin/env python3
"""
NetSuite System Discovery — Opiniion
Catalogs chart of accounts, departments, classes, subsidiaries, and custom records.
Outputs a formatted Data Dictionary Excel workbook.

Usage:
    python3 discover_netsuite.py
    python3 discover_netsuite.py --output ~/Desktop/NS_Discovery.xlsx
"""

import argparse
import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))
from connectors.netsuite_connect import NetSuiteClient

TEAL = "4FCBC5"
DARK_GRAY = "4C4D4E"
HEADER_GRAY = "F3F3F3"
BRAND_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
VALUE_FONT = Font(name="Calibri", size=10, color=DARK_GRAY)


def discover_chart_of_accounts(client: NetSuiteClient) -> pd.DataFrame:
    """Pull the full chart of accounts."""
    sql = """
    SELECT
        a.acctnumber AS account_number,
        a.acctname AS account_name,
        a.accttype AS account_type,
        a.generalrate AS general_rate,
        a.isinactive AS is_inactive,
        a.parent AS parent_id,
        a.description AS description
    FROM account a
    ORDER BY a.acctnumber
    """
    print("[Discovery] Pulling chart of accounts...")
    return client.query(sql)


def discover_departments(client: NetSuiteClient) -> pd.DataFrame:
    """Pull all departments."""
    sql = """
    SELECT
        d.id AS department_id,
        d.name AS department_name,
        d.parent AS parent_id,
        d.isinactive AS is_inactive
    FROM department d
    ORDER BY d.name
    """
    print("[Discovery] Pulling departments...")
    return client.query(sql)


def discover_classes(client: NetSuiteClient) -> pd.DataFrame:
    """Pull all classes."""
    sql = """
    SELECT
        c.id AS class_id,
        c.name AS class_name,
        c.parent AS parent_id,
        c.isinactive AS is_inactive
    FROM classification c
    ORDER BY c.name
    """
    print("[Discovery] Pulling classes...")
    return client.query(sql)


def discover_subsidiaries(client: NetSuiteClient) -> pd.DataFrame:
    """Pull all subsidiaries."""
    sql = """
    SELECT
        s.id AS subsidiary_id,
        s.name AS subsidiary_name,
        s.parent AS parent_id,
        s.isinactive AS is_inactive,
        s.country AS country
    FROM subsidiary s
    ORDER BY s.name
    """
    print("[Discovery] Pulling subsidiaries...")
    return client.query(sql)


def discover_transaction_types(client: NetSuiteClient) -> pd.DataFrame:
    """Pull distinct transaction types with record counts."""
    sql = """
    SELECT
        t.type AS transaction_type,
        COUNT(*) AS record_count,
        MIN(t.trandate) AS earliest_date,
        MAX(t.trandate) AS latest_date
    FROM transaction t
    GROUP BY t.type
    ORDER BY record_count DESC
    """
    print("[Discovery] Pulling transaction type summary...")
    return client.query(sql)


def discover_locations(client: NetSuiteClient) -> pd.DataFrame:
    """Pull all locations."""
    sql = """
    SELECT
        l.id AS location_id,
        l.name AS location_name,
        l.parent AS parent_id,
        l.isinactive AS is_inactive
    FROM location l
    ORDER BY l.name
    """
    print("[Discovery] Pulling locations...")
    return client.query(sql)


def write_discovery_workbook(
    coa: pd.DataFrame,
    departments: pd.DataFrame,
    classes: pd.DataFrame,
    subsidiaries: pd.DataFrame,
    txn_types: pd.DataFrame,
    locations: pd.DataFrame,
    output_path: str,
):
    """Write the NetSuite discovery workbook."""
    wb = Workbook()

    datasets = [
        ("Chart of Accounts", coa),
        ("Departments", departments),
        ("Classes", classes),
        ("Subsidiaries", subsidiaries),
        ("Transaction Types", txn_types),
        ("Locations", locations),
    ]

    for idx, (tab_name, df) in enumerate(datasets):
        if idx == 0:
            ws = wb.active
            ws.title = tab_name
        else:
            ws = wb.create_sheet(tab_name)

        ws.sheet_view.showGridLines = False
        ws["A1"] = "opiniion"
        ws["A1"].font = BRAND_FONT

        if df.empty:
            ws.cell(row=3, column=1, value="No data returned. Check query permissions.").font = VALUE_FONT
            continue

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 4, column=col_idx, value=value)
                cell.font = VALUE_FONT

        for col_idx in range(1, len(df.columns) + 1):
            max_len = max(
                len(str(df.columns[col_idx - 1])),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if not df.empty else 0,
            )
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A4"

    summary = wb.create_sheet("Summary", 0)
    summary.sheet_view.showGridLines = False
    summary["A1"] = "opiniion"
    summary["A1"].font = BRAND_FONT
    summary.cell(row=3, column=1, value="NetSuite Data Discovery Summary").font = Font(
        name="Calibri", size=14, bold=True, color=DARK_GRAY
    )
    summary.cell(row=4, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d %H:%M')}").font = VALUE_FONT

    row = 6
    for tab_name, df in datasets:
        summary.cell(row=row, column=1, value=tab_name).font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        summary.cell(row=row, column=2, value=f"{len(df)} records").font = VALUE_FONT
        row += 1

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 15

    wb.save(output_path)
    print(f"[Discovery] NetSuite workbook saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="NetSuite System Discovery")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    output_path = args.output or os.path.expanduser(
        f"~/Desktop/Opiniion_NetSuite_Discovery_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    )

    client = NetSuiteClient()
    print("[Discovery] Testing NetSuite connection...")
    if not client.test_connection():
        print("[Discovery] Connection failed. Check .env credentials.")
        sys.exit(1)

    coa = discover_chart_of_accounts(client)
    departments = discover_departments(client)
    classes = discover_classes(client)
    subsidiaries = discover_subsidiaries(client)
    txn_types = discover_transaction_types(client)
    locations = discover_locations(client)

    write_discovery_workbook(coa, departments, classes, subsidiaries, txn_types, locations, output_path)
    print("[Discovery] Done.")


if __name__ == "__main__":
    main()
